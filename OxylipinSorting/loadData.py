from openpyxl import load_workbook,Workbook
from sort import sortOxylipins
import re

# This file loads the passerini TGRL data, and creates excel worksheets that are passed to sortOxylipin function and saves the result

wb1 		=	load_workbook(filename = '../MetData/Passerini TGRL-Plasma Final Res.xlsx')
plasma 		=	wb1['Plasma Combined Lipid Results ']
tgrl 		=	wb1['TGRL Combined Lipid Results ']

results		=	Workbook()
pSort 		= 	results.active
pSort.title	=	'Plasma'
pSort.append(['Oxylipin','Parent FA','Type','D229','D360','D384','D391','D75','D274','D34','D356','D368','D370'])
print ('pSort starting...')
sortOxylipins(plasma,pSort,0)
print ('pSort done!')


tSort 		= 	results.create_sheet(title="TGRL")
tSort.append(['Oxylipin','Parent FA','Type','D229','D360','D384','D391','D75','D274','D34','D356','D368','D370'])

print ('tSort starting...')
sortOxylipins(tgrl,tSort,1)
print ('tSort done!')


results.save(filename='../results/sortedOxylipins_SA.xlsx')

