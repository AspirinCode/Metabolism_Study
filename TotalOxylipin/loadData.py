from openpyxl import load_workbook,Workbook
import numpy as np
import re
import scipy.stats as stat
import sort2 as sort


#Get excel sheet of sorted oxylipins
wb1 		= 	load_workbook(filename = '../results/sortedAll_SA.xlsx')
plasmawb 	=	wb1['Plasma']
tgrlwb 		=	wb1['TGRL']

plasma 		= 	[]
tgrl 		=	[]


#turns excel sheet in to array
for row in plasmawb.iter_rows(min_row=2):
	rowV = []

	for cell in row:
		rowV.append(cell.value)
	plasma.append(rowV)

for row in tgrlwb.iter_rows(min_row=2):
	rowV = []
	for cell in row:
		rowV.append(cell.value)
	tgrl.append(rowV)

#Create excel workbook to input data with plasma and tgrl sheets
results		=	Workbook()
pSort 		= 	results.active
pSort.title	=	'Plasma'
tSort 		= 	results.create_sheet(title="TGRL")


#run sorting and write to sheets
pRes = sort.sort(plasma,pSort)
tRes = sort.sort(tgrl,tSort)
#sort.write_results(pRes,pSort,pSum)
#sort.write_results(tRes,tSort,tSum)

#save
results.save(filename='../results/classify_all.xlsx')






