import openpyxl
import csv

wb1 = openpyxl.load_workbook('Passerini TGRL-Plasma Final Res.xlsx')
sheets = wb1.get_sheet_names()
plasma = wb1.get_sheet_by_name(sheets[2])

#wb2 = openpyxl.load_workbook('formatted.xlsx')
nefapp = []
nefaf = []
tofapp = []
tofaf = []

x=[]


for row in range(2, plasma.max_row + 1):
	A = plasma['A' + str(row)].value.encode('utf-8')
	B = plasma['B' + str(row)].value
	C = plasma['C' + str(row)].value
	D = plasma['D' + str(row)].value
	E = plasma['E' + str(row)].value
	F = plasma['F' + str(row)].value
	G = plasma['G' + str(row)].value
	H = plasma['H' + str(row)].value
	nefa = 0
	PP = 0
	if (A.find('%') != -1):
		E = E*100
		F = F*100
		G = G*100
		H = H*100

	if (abs(C)>0.548 or D<0.05):
		if ((A.lower().find('nefa') != -1) or (A.lower().find('ne') != -1)):
			nefa = 1;
		if (A.find('PP') != -1):
			PP = 1;

		if PP:
			if nefa:
				nefapp.append([A,B,'','',C,D,E,F,G,H])
			else:
				tofapp.append([A,B,'','',C,D,E,F,G,H])
		else:
			if nefa:
				nefaf.append([A,B,'','',C,D,E,F,G,H])
			else:
				tofaf.append([A,B,'','',C,D,E,F,G,H])

with open('Psum.csv', 'wb') as csvfile:
	writer = csv.writer(csvfile)
	writer.writerow(['Oxylipin','Name','Parent FA','Pathway','Correlation','P-Value','Anti-Athero Mean','Pro-Athero Mean', 'Anti- St Dev', 'Pro -St Dev'])
	writer.writerow(['Nefa-PP'])
	writer.writerows(nefapp)
	writer.writerow(' ')
	writer.writerow(['Nefa-F'])
	writer.writerows(nefaf)
	writer.writerow(' ')
	writer.writerow(['Tofa-PP'])
	writer.writerows(tofapp)
	writer.writerow(' ')
	writer.writerow(['Tofa-F'])
	writer.writerows(tofaf)	



