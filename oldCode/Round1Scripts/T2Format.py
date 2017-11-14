import openpyxl
import csv

wb1 = openpyxl.load_workbook('Passerini TGRL-Plasma Final Res.xlsx')
sheets = wb1.get_sheet_names()
plasma = wb1.get_sheet_by_name(sheets[4])

#wb2 = openpyxl.load_workbook('formatted.xlsx')
oxylipins = ['Alcohol','Ketone','Diol','Triol','Epoxide','Prostaglandin','Alcohol Precursor']
nefapp ={'Alcohol':[],'Ketone':[],'Diol':[],'Triol':[],'Epoxide':[],'Prostaglandin':[],'Alcohol Precursor':[]}
nefaf = {'Alcohol':[],'Ketone':[],'Diol':[],'Triol':[],'Epoxide':[],'Prostaglandin':[],'Alcohol Precursor':[]}
tofapp = {'Alcohol':[],'Ketone':[],'Diol':[],'Triol':[],'Epoxide':[],'Prostaglandin':[],'Alcohol Precursor':[]}
tofaf = {'Alcohol':[],'Ketone':[],'Diol':[],'Triol':[],'Epoxide':[],'Prostaglandin':[],'Alcohol Precursor':[]}

x=[]


for row in range(2, plasma.max_row + 1):

	A = plasma['A' + str(row)].value
	a = A.encode('utf-8')
	B = plasma['B' + str(row)].value
	C = plasma['C' + str(row)].value
	D = plasma['D' + str(row)].value
	E = plasma['E' + str(row)].value
	F = plasma['F' + str(row)].value
	G = plasma['G' + str(row)].value
	H = plasma['H' + str(row)].value
	I = plasma['I' + str(row)].value
	J = plasma['J' + str(row)].value
	nefa = 0
	PP = 0
	Type=[0,0,0,0,0,0,0]

	if (a.find('%') != -1):
		E = E*100
		F = F*100
		G = G*100
		H = H*100

	if ((a.lower().find('nefa') != -1) or (a.lower().find('ne') != -1)):
		nefa = 1;
	if (a.find('PP') != -1):
		PP = 1;
	if(I):	
		if (I.lower().find('alcohol')!= -1):
			Type[0]=1
		if (I.lower().find('ketone')!= -1):
			Type[1]=1			
		if (I.lower().find('diol')!= -1):
			Type[2]=1		
		if (I.lower().find('triol')!= -1):
			Type[3]=1
		if (I.lower().find('epoxide')!= -1):
			Type[4]=1
		if (I.lower().find('prostaglandin')!= -1):
			Type[5]=1
		if (I.lower().find('precursor')!= -1):
			Type[6]=1
			Type[0]=0

	for x in oxylipins:	
		if Type[oxylipins.index(x)]:
			if(nefa):
				if(PP):
					nefapp[x].append([A,B,C,D,E,F,G,H,I,J])
				else:
					nefaf[x].append([A,B,C,D,E,F,G,H,I,J])
			else:
				if(PP):
					tofapp[x].append([A,B,C,D,E,F,G,H,I,J])
				else:
					tofaf[x].append([A,B,C,D,E,F,G,H,I,J])

with open('T2sum.csv', 'wb') as csvfile:
	writer = csv.writer(csvfile)
	writer.writerow(['Oxylipin','Name','Correlation','P-Value','Anti-Athero Mean','Pro-Athero Mean', 'Anti- St Dev', 'Pro -St Dev','Type','Parent Fatty Acid'])
	writer.writerow(['NEFA-PP'])
	for x in oxylipins:
		writer.writerow([x])
		writer.writerows(nefapp[x])
		writer.writerows(' ')

	writer.writerow(['NEFA-F'])
	for x in oxylipins:
		writer.writerow([x])
		writer.writerows(nefaf[x])
		writer.writerows(' ')

	writer.writerow(['TOFA-PP'])
	for x in oxylipins:
		writer.writerow([x])
		writer.writerows(tofapp[x])
		writer.writerows(' ')

	writer.writerow(['TOFA-F'])
	for x in oxylipins:
		writer.writerow([x])
		writer.writerows(tofaf[x])
		writer.writerows(' ')

